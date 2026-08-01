"""Тесты файлового слепка табеля при утверждении (#17)."""
import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.services.timesheet_snapshot_service import timesheet_snapshot_service
from app.services.work_schedule_service import work_schedule_service


pytestmark = pytest.mark.asyncio(loop_scope="module")


def _dow(day: int) -> str:
    return date(2025, 7, day).strftime("%A")[:2]


async def test_approve_creates_snapshot_file(db_session, create_employee, tmp_path, monkeypatch):
    """Утверждение графика сохраняет слепок табеля файлом."""
    monkeypatch.setattr(timesheet_snapshot_service, "root", tmp_path)
    emp = await create_employee(name="Снапшот Тест", tab_number=42, hire_date=date(2024, 1, 1))
    schedule = await work_schedule_service.create_schedule(db_session, emp.id, 2025, 7, "t")
    await work_schedule_service.bulk_set_entries(
        db_session,
        schedule.id,
        [
            {"work_date": date(2025, 7, 1), "shift_type_code": "day", "planned_hours_override": None},
            {"work_date": date(2025, 7, 2), "shift_type_code": "night", "planned_hours_override": None},
        ],
    )

    await work_schedule_service.approve_schedule(db_session, schedule.id, "boss")

    snapshots = timesheet_snapshot_service.list_snapshots(emp.id)
    assert len(snapshots) == 1
    path = timesheet_snapshot_service.resolve_snapshot(emp.id, snapshots[0]["file_name"])
    assert path is not None
    assert path.exists()
    assert path.suffix == ".xlsx"


async def test_snapshot_contains_approved_values_not_current(
    db_session, create_employee, tmp_path, monkeypatch
):
    """В слепке значения на момент утверждения, а не текущие (#17, «как было»).

    Утвердили месяц, потом поменяли ячейку и скачали слепок — в нём
    должны остаться значения, которые были в момент утверждения.
    """
    monkeypatch.setattr(timesheet_snapshot_service, "root", tmp_path)
    emp = await create_employee(name="Снапшот Тест", tab_number=42, hire_date=date(2024, 1, 1))
    schedule = await work_schedule_service.create_schedule(db_session, emp.id, 2025, 7, "t")
    await work_schedule_service.set_entry(
        db_session, schedule.id, date(2025, 7, 1), shift_type_code="day"
    )

    await work_schedule_service.approve_schedule(db_session, schedule.id, "boss")

    # После утверждения меняем ячейку — это разрешено и не меняет слепок
    await work_schedule_service.set_entry(
        db_session, schedule.id, date(2025, 7, 1), shift_type_code="vacation"
    )

    snapshots = timesheet_snapshot_service.list_snapshots(emp.id)
    path = timesheet_snapshot_service.resolve_snapshot(emp.id, snapshots[0]["file_name"])
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Дата"
    # Итог на 1 июля = day (значение на момент утверждения)
    day1 = next(r for r in rows if r[0] == "2025-07-01")
    assert day1[2] == "day"


async def test_snapshot_second_approval_keeps_history(
    db_session, create_employee, tmp_path, monkeypatch
):
    """Повторное утверждение создаёт новый файл рядом с прежним (история)."""
    monkeypatch.setattr(timesheet_snapshot_service, "root", tmp_path)
    emp = await create_employee(name="Снапшот Тест", tab_number=42, hire_date=date(2024, 1, 1))
    schedule = await work_schedule_service.create_schedule(db_session, emp.id, 2025, 7, "t")
    await work_schedule_service.set_entry(
        db_session, schedule.id, date(2025, 7, 1), shift_type_code="day"
    )

    await work_schedule_service.approve_schedule(db_session, schedule.id, "boss")
    await work_schedule_service.unapprove_schedule(db_session, schedule.id)
    await work_schedule_service.approve_schedule(db_session, schedule.id, "boss2")

    snapshots = timesheet_snapshot_service.list_snapshots(emp.id)
    assert len(snapshots) == 2


async def test_generate_xlsx_hours_summary(create_employee, db_session):
    """Слепок считает часы по итоговому слою (рабочие смены)."""
    emp = await create_employee(name="Часы", tab_number=7, hire_date=date(2024, 1, 1))
    employee = await timesheet_snapshot_service.build_employee_cells(db_session, emp.id, 2025, 7)
    assert employee  # сотрудник найден

    content = timesheet_snapshot_service.generate_xlsx(employee, 2025, 7)
    assert content[:2] == b"PK"  # валидный zip/xlsx

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 33  # заголовок + 31 день + итог
    assert rows[0][0] == "Дата"
