"""Файловый слепок табеля при утверждении графика (#17).

Решение по #13: утверждение не блокирует правки, а сохраняет копию данных
файлом. На каждое утверждение генерируется отдельный .xlsx (история слепков),
чтобы можно было посмотреть «как было» на момент утверждения.

В слепок попадает итог за каждый день («ручное, иначе авто») по трёхслойной
модели ячейки, плюс часы по итоговому слою. Формат — Excel через openpyxl.
"""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.logging import get_audit_logger
from app.core.shift_types import get_shift_type
from app.services.timesheet_service import timesheet_import_service

logger = get_audit_logger()


def _safe_filename_part(value: str) -> str:
    return re.sub(r"[^0-9A-Za-zа-яА-ЯёЁ_-]+", "_", value).strip("_")[:40] or "employee"


class TimesheetSnapshotService:
    """Генерация и чтение файловых слепков табеля."""

    def __init__(self) -> None:
        self.root = Path(settings.TIMESHEET_SNAPSHOTS_PATH)

    def _employee_dir(self, employee_id: int) -> Path:
        return self.root / str(employee_id)

    async def build_employee_cells(
        self, db: AsyncSession, employee_id: int, year: int, month: int
    ) -> dict:
        """Собирает трёхслойные ячейки месяца для одного сотрудника.

        Переиспользует get_timesheet — то же правило «ручное, иначе авто»,
        что и в живой сетке, чтобы слепок не расходился с интерфейсом.
        """
        period_start = date(year, month, 1)
        period_end = date(year, month, 1).replace(day=28) + timedelta(days=4)
        period_end = period_end.replace(day=1) - timedelta(days=1)
        data = await timesheet_import_service.get_timesheet(
            db, period_start, period_end, employee_ids=[employee_id]
        )
        for emp in data.get("employees", []):
            if emp["id"] == employee_id:
                return emp
        return {}

    def _shift_label(self, code: Optional[str]) -> str:
        st = get_shift_type(code)
        return st.name if st else (code or "")

    def generate_xlsx(self, employee: dict, year: int, month: int) -> bytes:
        """Строит .xlsx: по строке на день месяца, итог из трёх слоёв.

        Колонки: Дата | День недели | Итог (код) | Итог (название) | Часы
        | Авто | Ручное. Итог = ручное, иначе авто (модель ячейки).
        """
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Табель"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        center = Alignment(horizontal="center", vertical="center")

        headers = ["Дата", "День", "Итог", "Смена", "Часы", "Авто", "Ручное"]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        days_in_month = (
            date(year, month, 1).replace(day=28) + timedelta(days=4)
        ).replace(day=1) - timedelta(days=1)
        dow_short = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        cells = employee.get("cells") or {}
        total_hours = 0.0

        row = 2
        current = date(year, month, 1)
        while current <= days_in_month:
            date_str = current.isoformat()
            cell = cells.get(date_str) or {}
            result_code = cell.get("result")
            auto_code = (cell.get("auto") or {}).get("shift_type_code")
            manual_code = (cell.get("manual") or {}).get("shift_type_code")
            hours = 0.0
            if result_code:
                st = get_shift_type(result_code)
                if st and st.is_working:
                    override = (cell.get("manual") or {}).get("planned_hours_override")
                    hours = float(override) if override is not None else st.planned_hours
                    total_hours += hours

            ws.cell(row=row, column=1, value=current.isoformat())
            ws.cell(row=row, column=2, value=dow_short[current.weekday()])
            ws.cell(row=row, column=3, value=result_code or "")
            ws.cell(row=row, column=4, value=self._shift_label(result_code))
            ws.cell(row=row, column=5, value=hours if hours else "")
            ws.cell(row=row, column=6, value=auto_code or "")
            ws.cell(row=row, column=7, value=manual_code or "")
            row += 1
            current += timedelta(days=1)

        ws.cell(row=row, column=4, value="Итого часов").font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_hours).font = Font(bold=True)

        for col, width in enumerate([12, 8, 10, 26, 8, 10, 10], start=1):
            ws.column_dimensions[chr(64 + col)].width = width

        import io

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def save_snapshot(
        self,
        employee: dict,
        year: int,
        month: int,
        employee_id: int,
        approved_by: str,
    ) -> Path:
        """Пишет .xlsx в каталог слепков сотрудника (история, не замена)."""
        content = self.generate_xlsx(employee, year, month)
        emp_name = _safe_filename_part(employee.get("name") or str(employee_id))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        snap_dir = self._employee_dir(employee_id)
        snap_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"{year}-{month:02d}_{ts}_{emp_name}.xlsx"
        file_path = snap_dir / file_name
        file_path.write_bytes(content)
        logger.info(
            "TIMESHEET SNAPSHOT",
            extra={
                "action": "timesheet_snapshot_create",
                "user_id": approved_by,
                "details": {"employee_id": employee_id, "file": file_name},
            },
        )
        return file_path

    def list_snapshots(self, employee_id: int) -> List[dict]:
        """Список слепков сотрудника, новые первыми."""
        snap_dir = self._employee_dir(employee_id)
        if not snap_dir.exists():
            return []
        items = []
        for p in sorted(snap_dir.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True):
            items.append(
                {
                    "file_name": p.name,
                    "size": p.stat().st_size,
                    "created_at": datetime.fromtimestamp(p.stat().st_mtime).isoformat(),
                }
            )
        return items

    def resolve_snapshot(self, employee_id: int, file_name: str) -> Optional[Path]:
        """Безопасное разрешение имени файла слепка (без traversal)."""
        if not file_name or Path(file_name).name != file_name or not file_name.endswith(".xlsx"):
            return None
        candidate = self._employee_dir(employee_id) / file_name
        if candidate.exists() and candidate.is_file():
            return candidate
        return None


timesheet_snapshot_service = TimesheetSnapshotService()
